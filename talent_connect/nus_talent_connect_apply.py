"""
NUS Talent Connect — Apply Helper
=================================
Opens each unvisited eligible job link and applies automatically.

Workflow:
  1. Load links from the filter output file (Excel or CSV).
  2. Skip links already marked as visited.
  3. Open each link in the browser.
  4. Auto-detect if already applied (badge "Applied" visible) → skip & mark visited.
  5. Auto-click the "Apply" button if present.
  6. Pause for user to complete the application form, then press Enter.
  7. Mark link as visited and persist.

Usage:
  python nus_talent_connect_apply.py
  python nus_talent_connect_apply.py --input temp/nus_internships_filtered.xlsx
  python nus_talent_connect_apply.py --reset
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import json
from pathlib import Path

import pyrootutils

root = pyrootutils.setup_root(__file__, dotenv=True, pythonpath=True, cwd=True)

from openpyxl import load_workbook

from tools import BrowserTools, async_request_user_interaction

SAVE_DIR = Path(__file__).parent / "temp"
DEFAULT_INPUT = SAVE_DIR / "nus_internships_filtered.xlsx"
DEFAULT_CSV_INPUT = SAVE_DIR / "nus_internships_filtered.csv"
VISITED_FILE = SAVE_DIR / "nus_jobs_visited.json"

# --- Selectors discovered via exploration ---
# Applied badge: <span class="badge-small badge-success">Applied</span>
APPLIED_BADGE_SELECTOR = "span.badge-small.badge-success"
# Apply button: <button class="btn btn_primary hide-sm">Apply</button>
APPLY_BUTTON_SELECTOR = "button.btn.btn_primary"


def _load_json(path: Path) -> dict | list | None:
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None


def _save_json(path: Path, data: dict | list) -> None:
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def _normalize_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y", "t"}


def _load_rows_from_excel(path: Path) -> list[dict]:
    wb = load_workbook(path)
    sheet = wb.active
    assert sheet is not None
    headers = []
    rows: list[dict] = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(cell).strip() if cell is not None else "" for cell in row]
            continue
        data = {headers[j]: row[j] for j in range(len(headers))}
        rows.append(data)
    return rows


def _load_rows_from_csv(path: Path) -> list[dict]:
    rows: list[dict] = []
    with path.open("r", newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            rows.append(row)
    return rows


def _load_rows(path: Path) -> list[dict]:
    if path.suffix.lower() == ".xlsx":
        return _load_rows_from_excel(path)
    if path.suffix.lower() == ".csv":
        return _load_rows_from_csv(path)
    raise ValueError(f"Unsupported input format: {path.suffix}")


def _get_links(rows: list[dict]) -> list[str]:
    links = []
    for row in rows:
        link = (row.get("link") or "").strip()
        if not link:
            continue
        eligible = _normalize_bool(row.get("eligible"))
        if eligible:
            links.append(link)
    return links


def _load_visited() -> set[str]:
    data = _load_json(VISITED_FILE)
    if isinstance(data, list):
        return {str(x) for x in data}
    if isinstance(data, dict) and "visited" in data:
        return {str(x) for x in data.get("visited", [])}
    return set()


def _save_visited(visited: set[str]) -> None:
    _save_json(VISITED_FILE, sorted(visited))


async def _is_already_applied(bt: BrowserTools) -> bool:
    """Check if the current job page shows an 'Applied' badge."""
    try:
        count = await bt.query_selector_all(APPLIED_BADGE_SELECTOR)
        return count > 0
    except Exception:
        return False


async def _click_apply(bt: BrowserTools) -> bool:
    """Click the 'Apply' button on the job detail page.

    Returns True if a button was found and clicked, False otherwise.
    """
    try:
        count = await bt.query_selector_all(APPLY_BUTTON_SELECTOR)
        if count == 0:
            return False
        await bt.click(APPLY_BUTTON_SELECTOR, timeout=5000)
        return True
    except Exception as exc:
        print(f"  [warn] Could not click Apply: {exc}")
        return False


async def main() -> None:
    parser = argparse.ArgumentParser(description="NUS Talent Connect — Apply Helper")
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help="Path to the filtered output file (.xlsx or .csv).",
    )
    parser.add_argument(
        "--reset",
        action="store_true",
        help="Clear visited cache before starting.",
    )
    args = parser.parse_args()

    input_path = args.input
    if not input_path.exists():
        if input_path == DEFAULT_INPUT and DEFAULT_CSV_INPUT.exists():
            input_path = DEFAULT_CSV_INPUT
        else:
            raise FileNotFoundError(f"Input file not found: {input_path}")

    rows = _load_rows(input_path)
    links = _get_links(rows)
    if not links:
        print("No eligible links found in the input file.")
        return

    if args.reset and VISITED_FILE.exists():
        VISITED_FILE.unlink()

    visited = _load_visited()
    pending = [link for link in links if link not in visited]

    print(f"Eligible links: {len(links)}")
    print(f"Already visited: {len(visited)}")
    print(f"Pending: {len(pending)}")

    if not pending:
        print("Nothing to open. All eligible links are visited.")
        return

    bt = BrowserTools(headless=False)
    await bt.start(site_name="nus_talent_connect")

    stats = {"applied": 0, "already_applied": 0, "skipped": 0}

    try:
        for idx, link in enumerate(pending, start=1):
            print(f"\n[{idx}/{len(pending)}] Opening: {link}")
            await bt.navigate(link)
            await bt.wait_for_timeout(2000)

            # --- Auto-detect already applied ---
            if await _is_already_applied(bt):
                print("  Already applied — skipping.")
                visited.add(link)
                _save_visited(visited)
                stats["already_applied"] += 1
                continue

            # --- Auto-click Apply button ---
            clicked = await _click_apply(bt)
            if clicked:
                print("  Clicked Apply button.")
                await bt.wait_for_timeout(1000)
            else:
                print("  [warn] No Apply button found on this page.")
                stats["skipped"] += 1

            # --- Wait for user to complete application form ---
            await async_request_user_interaction(
                "Complete the application form (or just review), then press Enter to continue."
            )

            visited.add(link)
            _save_visited(visited)
            stats["applied"] += 1
            print("  Marked visited.")
    finally:
        await bt.stop()

    print(
        f"\nDone. Applied: {stats['applied']}, "
        f"Already applied: {stats['already_applied']}, "
        f"Skipped: {stats['skipped']}"
    )


if __name__ == "__main__":
    asyncio.run(main())
